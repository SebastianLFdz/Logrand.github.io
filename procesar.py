import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os
import difflib

def procesar_archivos(plantilla_path, datos_path, output_folder):
    df_altas = pd.read_excel(datos_path, sheet_name="Altas")
    df_bajas = pd.read_excel(datos_path, sheet_name="Bajas")

    # Normalizamos encabezados
    df_altas.columns = df_altas.columns.str.strip()
    df_bajas.columns = df_bajas.columns.str.strip()

    # Correlación salas
    sala_correspondencia = {
        "JUBILEE": "Jubilee",
        "NEW YORK": "New York",
        "PARADISE": "Paradise",
        "VIVA MEXICO": "Viva Mexico",
        "HOLLYWOOD CONSTITUCION": "HWC",
        "HOLLYWOOD VALLE ALTO": "HWVA",
        "VIVENTO APODACA": "Vivento Apodaca",
        "JUBILEE CANCUN": "Jubilee Cancún",
        "GOLDEN ISLAND": "Golden Island",
        "GRAND LEON": "Grand León",
        "TAJ MAHAL": "Taj Mahal",
        "JUBILEE CDMX": "Jubilee CDMX",
        "VIVENTO CULIACAN": "Vivento Culiacan",
        "VIVENTO ZAPOPAN": "Vivento Zapopan",
        "EL DORADO": "El Dorado"
    }

    salas_orden = ["Jubilee", "New York", "Paradise", "Viva Mexico", "HWC", "HWVA", "Vivento Apodaca", 
                   "Jubilee Cancún", "Golden Island", "Grand León", "Taj Mahal", "Jubilee CDMX", 
                   "Vivento Culiacan", "Vivento Zapopan", "El Dorado"]

    posiciones_salas = {
        sala: chr(67 + i) for i, sala in enumerate(salas_orden)
    }

    posiciones_filas = {
        "Altas Generales": 10,
        "Bajas Generales": 11
    }

    wb = load_workbook(plantilla_path)
    ws = wb.active

    # ALTAS Y BAJAS GENERALES
    conteo_altas = df_altas["Sala AP"].map(lambda s: sala_correspondencia.get(s.strip().upper(), None)).value_counts()
    conteo_bajas = df_bajas["Sala AP"].map(lambda s: sala_correspondencia.get(s.strip().upper(), None)).value_counts()

    for sala in salas_orden:
        columna = posiciones_salas[sala]

        ws[f"{columna}{posiciones_filas['Altas Generales']}"] = int(conteo_altas.get(sala, 0))
        ws[f"{columna}{posiciones_filas['Bajas Generales']}"] = int(conteo_bajas.get(sala, 0))

    # ======================== Procesamiento de puestos ======================

    puestos_info = {
        "Asistente de Servicio": (17, 18),
        "Mesero Sala": (24, 25),
        "Mesero Restaurante": (31, 32),
        "Cajero (a)": (38, 39),
        "Valet Parking": (45, 46),
        "Mac": (52, 53),
        "Cocinero A": (59, 60),
        "Cocinero B": (66, 67),
        "Ayudante de Cocina": (73, 74),
        "Lavaloza": (80, 81),
        "Supervisor de Porteros": (87, 88),
        "Portero Interno": (94, 95),
        "Portero I. Femenino": (101, 102),
        "Portero Externo": (108, 109),
        "Tecnico de Mantenimiento": (115, 116),
        "Coordinador de Imagen": (122, 123),
        "Imagen": (129, 130),
        "Supervisor Sportbar": (136, 137),
        "Asesor Sportbar": (143, 144),
        "Pit Boss": (150, 151),
        "Supervisor de Mesas": (157, 158),
        "Dealer": (164, 165)
    }

    puesto_mapeo = {
    "Asistente de Servicio": "Asistente de Servicio",
    "Mesero Sala": "Mesero Sala",
    "Mesero": "Mesero Sala",  # <- Unificado con Mesero Sala
    "Mesero Restaurante": "Mesero Restaurante",
    "Cajero (a)": "Cajero (a)",
    "Valet Parking": "Valet Parking",
    "Mac": "Mac",
    "Cocinero A": "Cocinero A",
    "Cocinero B": "Cocinero B",
    "Ayudante de Cocina": "Ayudante de Cocina",
    "Lavaloza": "Lavaloza",
    "Supervisor de Porteros": "Supervisor de Porteros",
    "Portero Interno": "Portero Interno",
    "Portero": "Portero Interno",  # Ya lo teníamos
    "Portero I. Femenino": "Portero I. Femenino",
    "Portero Externo": "Portero Externo",
    "Tecnico de Mantenimiento": "Tecnico de Mantenimiento",
    "Coordinador de Imagen": "Coordinador de Imagen",
    "Imagen": "Imagen",
    "Supervisor Sportbar": "Supervisor Sportbar",
    "Asesor Sportbar": "Asesor Sportbar",
    "Pit Boss": "Pit Boss",
    "Supervisor de Mesas": "Supervisor de Mesas",
    "Dealer": "Dealer"
    }

    # Inicializamos los conteos
    conteo_altas_puestos = {puesto: {sala: 0 for sala in salas_orden} for puesto in puestos_info}
    conteo_bajas_puestos = {puesto: {sala: 0 for sala in salas_orden} for puesto in puestos_info}

    for _, row in df_altas.iterrows():
        puesto_excel = row['Puesto OM']
        sala_excel = row['Sala AP']
        puesto_logico = puesto_mapeo.get(puesto_excel.strip(), None)
        sala_logica = sala_correspondencia.get(sala_excel.strip().upper(), None)

        if puesto_logico and sala_logica:
            conteo_altas_puestos[puesto_logico][sala_logica] += 1

    for _, row in df_bajas.iterrows():
        puesto_excel = row['Puesto OM']
        sala_excel = row['Sala AP']
        puesto_logico = puesto_mapeo.get(puesto_excel.strip(), None)
        sala_logica = sala_correspondencia.get(sala_excel.strip().upper(), None)

        if puesto_logico and sala_logica:
            conteo_bajas_puestos[puesto_logico][sala_logica] += 1

    # Escribimos resultados
    for puesto, (fila_alta, fila_baja) in puestos_info.items():
        for sala in salas_orden:
            columna = posiciones_salas[sala]
            ws[f"{columna}{fila_alta}"] = conteo_altas_puestos[puesto][sala]
            ws[f"{columna}{fila_baja}"] = conteo_bajas_puestos[puesto][sala]

    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    output_filename = f"Plantilla Ajustada ({fecha_actual}).xlsx"
    output_path = os.path.join(output_folder, output_filename)
    wb.save(output_path)

    for df_nombre, df in [("Altas", df_altas), ("Bajas", df_bajas)]:
        if "Puesto OM" not in df.columns:
            print(f"❌ ERROR: La columna 'Puesto OM' no se encontró en el DataFrame de {df_nombre}.")
            print(f"🧪 Columnas disponibles en {df_nombre}: {list(df.columns)}")
        else:
            df.columns = df.columns.str.strip()  # Sanear espacios ocultos
            puestos_en_excel = set(df["Puesto OM"].dropna().unique())
            puestos_mapeados = set(puesto_mapeo.keys())

            puestos_faltantes = puestos_en_excel - puestos_mapeados

            if puestos_faltantes:
                print(f"⚠️ Puestos no mapeados encontrados en {df_nombre}:")
                for puesto in puestos_faltantes:
                    sugerencias = difflib.get_close_matches(puesto, puesto_mapeo.keys(), n=1, cutoff=0.8)
                    sugerencia = f" ¿Quizás quisiste decir: {sugerencias[0]}?" if sugerencias else ""
                    print(f"  - {puesto}{sugerencia}")

    return output_filename
